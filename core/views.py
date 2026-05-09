from calendar import month
import math
import csv
from urllib import request

from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Count
from django.http import HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from datetime import timedelta
from django.contrib import messages
from django.core.mail import send_mail 
from django.http import HttpResponse
from collections import OrderedDict
from django.db.models.functions import TruncDate
from .models import Review
from django.db.models import Avg
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.hashers import make_password
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from openpyxl import Workbook
from django.db.models import Count, Sum
from django.db.models.functions import TruncDate
from django.utils import timezone
from datetime import timedelta
import openpyxl

from .models import (
    Category,
    DiningTable,
    Employee,
    FavoriteRestaurant,
    Reservation,
    Restaurant,
    Review,
    Profile,
    Product,
    WorkShift,
    Attendance,
    EmployeePerformance,
    TableReservation,
    RestaurantTable,
)

def profile(request):
    user = request.user

    profile_obj, created = Profile.objects.get_or_create(user=user)

    if request.method == "POST":
        full_name = request.POST.get("full_name")
        email = request.POST.get("email")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")
        avatar = request.FILES.get("avatar")

        if full_name:
            user.first_name = full_name

        if email:
            user.email = email

        if password:
            if password == confirm_password:
                user.password = make_password(password)

        user.save()

        if avatar:
            profile_obj.avatar = avatar
            profile_obj.save()

    return render(request, "profile.html")

# ======= REVIEW ĐÁNH GIÁ QUÁN ĂN =====#
@login_required
def add_review(request, restaurant_id):
    restaurant = get_object_or_404(Restaurant, id=restaurant_id, is_active=True)

    # chỉ người đã từng đặt bàn và đơn đã duyệt mới được đánh giá
    has_booked = Reservation.objects.filter(
        restaurant=restaurant,
        customer_name=request.user.username,
        status="confirmed"
    ).exists()

    if not has_booked:
        return HttpResponseForbidden("Bạn chỉ có thể đánh giá sau khi đã đặt bàn và được duyệt.")

    if request.method == "POST":
        rating = int(request.POST.get("rating", 5))
        comment = request.POST.get("comment", "").strip()

        if rating < 1 or rating > 5:
            rating = 5

        Review.objects.update_or_create(
            restaurant=restaurant,
            user=request.user,
            defaults={
                "rating": rating,
                "comment": comment,
            }
        )

        return redirect("restaurant_detail", restaurant_id=restaurant.id)

    return redirect("restaurant_detail", restaurant_id=restaurant.id)


@login_required
def delete_review(request, restaurant_id):
    restaurant = get_object_or_404(Restaurant, id=restaurant_id, is_active=True)
    review = Review.objects.filter(restaurant=restaurant, user=request.user).first()

    if review:
        review.delete()

    return redirect("restaurant_detail", restaurant_id=restaurant.id)

def restaurant_detail(request, restaurant_id):
    restaurant = get_object_or_404(Restaurant, id=restaurant_id, is_active=True)
    reviews = restaurant.reviews.select_related("user").all()

    avg_rating = 0
    if reviews.exists():
        avg_rating = round(sum(r.rating for r in reviews) / reviews.count(), 1)

    user_review = None
    has_booked = False

    if request.user.is_authenticated:
        user_review = reviews.filter(user=request.user).first()
        has_booked = Reservation.objects.filter(
            restaurant=restaurant,
            customer_name=request.user.username,
            status="confirmed"
        ).exists()

    return render(request, "restaurant_detail.html", {
        "restaurant": restaurant,
        "reviews": reviews,
        "avg_rating": avg_rating,
        "user_review": user_review,
        "has_booked": has_booked,
    })

@csrf_exempt
def add_restaurant_from_map(request):
    if request.method == "POST":
        name = request.POST.get("name")
        address = request.POST.get("address", "")
        lat = request.POST.get("latitude")
        lng = request.POST.get("longitude")
        description = request.POST.get("description", "")

        if not name or not lat or not lng:
            return JsonResponse({
                "success": False,
                "message": "Thiếu tên quán hoặc tọa độ."
            })

        restaurant = Restaurant.objects.create(
            name=name,
            address=address,
            latitude=lat,
            longitude=lng,
            description=description
        )

        return JsonResponse({
            "success": True,
            "id": restaurant.id,
            "name": restaurant.name,
            "lat": restaurant.latitude,
            "lng": restaurant.longitude
        })

    return JsonResponse({"success": False})

@csrf_exempt
def delete_restaurant(request, id):
    try:
        restaurant = Restaurant.objects.get(id=id)
        restaurant.delete()
        return JsonResponse({
            "success": True
        })
    except Restaurant.DoesNotExist:
        return JsonResponse({
            "success": False,
            "message": "Không tìm thấy quán"
        })

# ====== GỬI MAIL ====== #
def test_mail(request):
    send_mail(
        'Test đặt bàn',
        'Bạn đã đặt bàn thành công!',
        'from@example.com',
        ['test@gmail.com'],
        fail_silently=False,
    )
    return HttpResponse("Đã gửi mail!")

# ====== Haversine distance ======
def haversine_km(lat1, lng1, lat2, lng2):
    R = 6371.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lng2 - lng1)

    a = (
        math.sin(dphi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    )
    return 2 * R * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# ====== HOME ======
def home_map(request):
    categories = Category.objects.all()
    center = {"lat": 12.67, "lng": 108.05, "zoom": 13}
    return render(request, "home_map.html", {
            "categories": categories,
            "center": center,
        },
    )


# ====== GEOJSON ======
def restaurants_geojson(request):
    qs = Restaurant.objects.filter(is_active=True)

    q = request.GET.get("q")
    if q:
        qs = qs.filter(name__icontains=q)

    cat = request.GET.get("category")
    if cat and cat.isdigit():
        qs = qs.filter(category_id=int(cat))

    only_favorites = request.GET.get("only_favorites")
    if only_favorites == "1" and request.user.is_authenticated:
        favorite_ids = FavoriteRestaurant.objects.filter(
            user=request.user
        ).values_list("restaurant_id", flat=True)
        qs = qs.filter(id__in=favorite_ids)

    lat = request.GET.get("lat")
    lng = request.GET.get("lng")
    radius_km = request.GET.get("radius_km")

    if lat and lng and radius_km:
        try:
            u_lat = float(lat)
            u_lng = float(lng)
            r = float(radius_km)

            filtered = []
            for res in qs:
                reviews = res.reviews.all()
                avg_rating = 0
                review_count = reviews.count()

                if review_count > 0:
                    avg_rating = round(sum(r.rating for r in reviews) / review_count, 1)
                d = haversine_km(u_lat, u_lng, res.latitude, res.longitude)
                if d <= r:
                    filtered.append((res, d))

            filtered.sort(key=lambda x: x[1])
            qs = [x[0] for x in filtered]
        except Exception:
            pass

    favorite_ids = set()
    if request.user.is_authenticated:
        favorite_ids = set(
            FavoriteRestaurant.objects.filter(user=request.user).values_list(
                "restaurant_id", flat=True
            )
        )

    features = []

    for res in qs:
        normal_available = DiningTable.objects.filter(
            restaurant=res,
            table_type="normal",
            status="available",
        ).exists()

        vip_available = DiningTable.objects.filter(
            restaurant=res,
            table_type="vip",
            status="available",
        ).exists()

        reviews = res.reviews.all()
        avg_rating = 0
        review_count = reviews.count()

        if review_count > 0:
            avg_rating = round(sum(r.rating for r in reviews) / review_count, 1)

        features.append({
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": [res.longitude, res.latitude],
            },
            "properties": {
                "id": res.id,
                "name": res.name,
                "address": res.address,
                "phone": res.phone,
                "category": res.category.name if res.category else None,
                "price_level": res.price_level,
                "image_url": getattr(res, "image_url", ""),
                "normal_available": normal_available,
                "vip_available": vip_available,
                "table_status": "Còn bàn" if normal_available or vip_available else "Hết bàn",
                "favorited": res.id in favorite_ids,
                "avg_rating": avg_rating,
                "review_count": review_count,
            },
        })

    return JsonResponse({"type": "FeatureCollection", "features": features})


# ====== ĐẶT BÀN ======
def reserve(request):
    restaurants = Restaurant.objects.filter(is_active=True).order_by("name")
    selected_restaurant_id = request.GET.get("restaurant", "")

    if request.method == "POST":
        restaurant_id = request.POST.get("restaurant")
        customer_name = request.POST.get("customer_name", "").strip()
        customer_phone = request.POST.get("customer_phone", "").strip()
        guests = request.POST.get("guests")
        booking_time = request.POST.get("booking_time")
        note = request.POST.get("note", "").strip()
        booking_type = request.POST.get("booking_type", "normal")

        if not all([restaurant_id, customer_name, customer_phone, guests, booking_time]):
            return render(
                request,
                "reserve.html",
                {
                    "restaurants": restaurants,
                    "selected_restaurant_id": restaurant_id or "",
                    "error": "Vui lòng nhập đầy đủ thông tin!",
                },
            )

        restaurant = get_object_or_404(Restaurant, id=restaurant_id)

        try:
            dt = timezone.make_aware(timezone.datetime.fromisoformat(booking_time))
        except Exception:
            return render(
                request,
                "reserve.html",
                {
                    "restaurants": restaurants,
                    "selected_restaurant_id": restaurant_id,
                    "error": "Thời gian không hợp lệ.",
                },
            )

        if dt <= timezone.now():
            return render(
                request,
                "reserve.html",
                {
                    "restaurants": restaurants,
                    "selected_restaurant_id": restaurant_id,
                    "error": "Thời gian phải ở tương lai.",
                },
            )

        Reservation.objects.create(
            restaurant=restaurant,
            customer_name=customer_name,
            customer_phone=customer_phone,
            guests=int(guests),
            booking_time=dt,
            note=note,
            status="pending",
            booking_type=booking_type,
        )

        return redirect("reserve_success")

    return render(
        request,
        "reserve.html",
        {
            "restaurants": restaurants,
            "selected_restaurant_id": selected_restaurant_id,
        },
    )
    

def reserve_success(request):
    return render(request, "reserve_success.html")


# ====== STAFF ======
def is_staff_user(user):
    return user.is_authenticated and Employee.objects.filter(user=user).exists()


def staff_login_view(request):
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()

        user = authenticate(request, username=username, password=password)

        if user is not None and Employee.objects.filter(user=user).exists():
            login(request, user)
            return redirect("staff_dashboard")

        return render(
            request,
            "staff_login.html",
            {
                "error": "Sai tài khoản, mật khẩu hoặc đây không phải tài khoản nhân viên."
            },
        )

    return render(request, "staff_login.html")

@login_required
def staff_dashboard(request):
    if not is_staff_user(request.user):
        return HttpResponseForbidden("Bạn không phải nhân viên.")

    tables = DiningTable.objects.select_related("restaurant").order_by("restaurant__name", "code")
    reservations = Reservation.objects.select_related("restaurant").order_by("-created_at")
    restaurants = Restaurant.objects.filter(is_active=True).order_by("name")

    # Thống kê bàn
    total_tables = tables.count()
    available = tables.filter(status="available").count()
    reserved = tables.filter(status="reserved").count()
    occupied = tables.filter(status="occupied").count()
    cleaning = tables.filter(status="cleaning").count()

    # Bàn VIP / thường
    vip_tables = tables.filter(table_type="vip").count()
    normal_tables = tables.filter(table_type="normal").count()

    # Đơn chờ xử lý
    pending_reservations = reservations.filter(status="pending")

    # Thống kê theo nhà hàng
    restaurant_stats = []
    for restaurant in restaurants:
        restaurant_tables = tables.filter(restaurant=restaurant)
        restaurant_stats.append({
            "name": restaurant.name,
            "total_tables": restaurant_tables.count(),
            "available": restaurant_tables.filter(status="available").count(),
            "reserved": restaurant_tables.filter(status="reserved").count(),
            "occupied": restaurant_tables.filter(status="occupied").count(),
            "cleaning": restaurant_tables.filter(status="cleaning").count(),
        })

    # Thông báo mới
    notifications = []
    for r in pending_reservations[:5]:
        notifications.append(
            f"Đơn mới từ {r.customer_name} tại {r.restaurant.name}"
        )

    return render(request, "staff_dashboard.html", {
        "tables": tables,
        "reservations": reservations,
        "restaurants": restaurants,

        "total_tables": total_tables,
        "available": available,
        "reserved": reserved,
        "occupied": occupied,
        "cleaning": cleaning,

        "vip_tables": vip_tables,
        "normal_tables": normal_tables,

        "pending_reservations": pending_reservations,
        "restaurant_stats": restaurant_stats,
        "notifications": notifications,
    })

@login_required
def update_table_status(request, table_id):
    if not is_staff_user(request.user):
        return HttpResponseForbidden("Không có quyền")

    table = get_object_or_404(DiningTable, id=table_id)

    if request.method == "POST":
        status = request.POST.get("status")
        if status in {"available", "reserved", "occupied", "cleaning"}:
            table.status = status
            table.save()

    return redirect("staff_dashboard")


# ====== LOGIN ADMIN ======
def admin_login_view(request):
    if request.method == "POST":
        user = authenticate(
            request,
            username=request.POST.get("username"),
            password=request.POST.get("password"),
        )

        if user and (user.is_staff or user.is_superuser):
            login(request, user)
            return redirect("admin_dashboard")

        return render(request, "admin_login.html", {"error": "Sai tài khoản"})

    return render(request, "admin_login.html")


# ====== LOGIN USER ======
def user_login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(
            request,
            username=username,
            password=password,
        )

        if user is not None and not user.is_staff:
            login(request, user)
            return redirect("user_profile")

        return render(
            request,
            "user_login.html",
            {
                "error": "Sai tài khoản, mật khẩu hoặc đây không phải tài khoản khách."
            },
        )

    return render(request, "user_login.html")


# ====== REGISTER ======
def register_view(request):
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "").strip()
        confirm_password = request.POST.get("confirm_password", "").strip()

        if not username or not password:
            return render(
                request,
                "register.html",
                {"error": "Mật khẩu xác nhận không khớp."},
            )

        if User.objects.filter(username=username).exists():
            return render(
                request,
                "register.html",
                {"error": "Tên đăng nhập đã tồn tại."},
            )

        if email and User.objects.filter(email=email).exists():
            return render(
                request,
                "register.html",
                {"error": "Email đã được sử dụng."},
            )

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
        )
        login(request, user)
        return redirect("user_profile")

    return render(request, "register.html")


def logout_view(request):
    logout(request)
    return redirect("user_login")


# ====== ADMIN DUYỆT / TỪ CHỐI ======
@login_required
def approve_reservation(request, reservation_id):
    if request.method != "POST":
        return redirect("admin_dashboard")

    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    reservation = get_object_or_404(Reservation, id=reservation_id)

    if reservation.status != "pending":
        messages.warning(request, "Đơn này đã được xử lý trước đó.")
        return redirect("admin_dashboard")

    available_table = DiningTable.objects.filter(
        restaurant=reservation.restaurant,
        table_type=reservation.booking_type,
        status="available",
    ).first()

    if not available_table:
        messages.error(
            request,
            f"Không thể duyệt: quán {reservation.restaurant.name} đã hết bàn {reservation.get_booking_type_display()}."
        )
        return redirect("admin_dashboard")

    available_table.status = "reserved"
    available_table.save()

    reservation.status = "confirmed"
    reservation.save()

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"success": True, "status": "confirmed"})

    messages.success(request, "Duyệt đơn thành công.")
    return redirect("admin_dashboard")

@login_required
def reject_reservation(request, reservation_id):
    if request.method != "POST":
        return redirect("admin_dashboard")

    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    reservation = get_object_or_404(Reservation, id=reservation_id)

    if reservation.status == "pending":
        reservation.status = "cancelled"
        reservation.save()

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"success": True, "status": "cancelled"})

    return redirect("admin_dashboard")




# ====== ADMIN DASHBOARD ======
@login_required
def admin_dashboard(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    restaurants = Restaurant.objects.filter(is_active=True).order_by("name")
    reservations = Reservation.objects.select_related("restaurant").order_by("-created_at")
    tables = DiningTable.objects.select_related("restaurant").all()
    categories = Category.objects.all()

    total_restaurants = restaurants.count()
    total_categories = categories.count()
    total_tables = tables.count()
    total_recent_reservations = reservations.count()

    available_count = tables.filter(status="available").count()
    reserved_count = tables.filter(status="reserved").count()
    occupied_count = tables.filter(status="occupied").count()
    cleaning_count = tables.filter(status="cleaning").count()

    pending_reservations = reservations.filter(status="pending")[:6]
    latest_reservations = reservations[:5]
    latest_reservation = reservations.first()

    now = timezone.now()
    current_month = now.month
    current_year = now.year
    current_day = now.day

    monthly_items = Reservation.objects.filter(
        status="confirmed",
        booking_time__month=current_month,
        booking_time__year=current_year
    ).select_related("restaurant")

    yearly_items = Reservation.objects.filter(
        status="confirmed",
        booking_time__year=current_year
    ).select_related("restaurant")

    monthly_revenue = 0
    yearly_revenue = 0

    for item in monthly_items:
        monthly_revenue += item.guests * item.restaurant.price_level * 100000

    for item in yearly_items:
        yearly_revenue += item.guests * item.restaurant.price_level * 100000

    today_reservations = Reservation.objects.filter(
        booking_time__date=now.date()
    ).select_related("restaurant").order_by("booking_time")[:6]

    # Doanh thu 7 ngày gần nhất
    recent_confirmed = Reservation.objects.filter(
        status="confirmed"
    ).select_related("restaurant").order_by("booking_time")
 
    daily_revenue_map = OrderedDict()

    for item in recent_confirmed:
       day = item.booking_time.strftime("%d/%m")
       amount = item.guests * item.restaurant.price_level * 100000
       if day not in daily_revenue_map:
          daily_revenue_map[day] = 0
       daily_revenue_map[day] += amount

    # chỉ lấy 7 mốc gần nhất
    daily_labels = list(daily_revenue_map.keys())[-7:]
    daily_values = list(daily_revenue_map.values())[-7:]

    top_restaurants = (
    Restaurant.objects.filter(is_active=True)
    .annotate(total_bookings=Count("reservation"))
    .order_by("-total_bookings")[:5]
)


    return render(request, "admin_dashboard_custom.html", {
        "restaurants": restaurants,
        "reservations": reservations,
        "tables": tables,
        "categories": categories,
        "daily_labels": daily_labels,
        "daily_values": daily_values,
        "top_restaurants": top_restaurants,

        "total_restaurants": total_restaurants,
        "total_categories": total_categories,
        "total_tables": total_tables,
        "total_recent_reservations": total_recent_reservations,

        "available_count": available_count,
        "reserved_count": reserved_count,
        "occupied_count": occupied_count,
        "cleaning_count": cleaning_count,

        "pending_reservations": pending_reservations,
        "latest_reservations": latest_reservations,
        "latest_reservation": latest_reservation,
        "today_reservations": today_reservations,

        "monthly_revenue": monthly_revenue,
        "yearly_revenue": yearly_revenue,
        "current_month": current_month,
        "current_year": current_year,
        "current_day": current_day,
    })

def admin_dashboard_custom(request):
    return render(request, 'admin/dashboard.html')



# ====== USER PROFILE ======
@login_required
def user_profile(request):
    profile_obj, created = Profile.objects.get_or_create(user=request.user)

    my_reservations = Reservation.objects.filter(
        customer_name=request.user.username
    ).select_related("restaurant").order_by("-created_at")

    approved_count = my_reservations.filter(status="confirmed").count()
    rejected_count = my_reservations.filter(status="cancelled").count()
    pending_count = my_reservations.filter(status="pending").count()

    notifications = []
    for r in my_reservations[:5]:
        if r.status == "confirmed":
            notifications.append(f"Đơn đặt bàn tại {r.restaurant.name} đã được duyệt.")
        elif r.status == "cancelled":
            notifications.append(f"Đơn đặt bàn tại {r.restaurant.name} đã bị từ chối.")
        else:
            notifications.append(f"Đơn đặt bàn tại {r.restaurant.name} đang chờ duyệt.")

    success_message = ""
    error = ""

    if request.method == "POST":
        full_name = request.POST.get("full_name", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "").strip()
        confirm_password = request.POST.get("confirm_password", "").strip()
        avatar = request.FILES.get("avatar")

        if full_name:
            parts = full_name.split()
            if len(parts) == 1:
                request.user.first_name = parts[0]
                request.user.last_name = ""
            else:
                request.user.first_name = " ".join(parts[:-1])
                request.user.last_name = parts[-1]

        if email:
            request.user.email = email

        if password:
            if password != confirm_password:
                error = "Mật khẩu xác nhận không khớp."
            else:
                request.user.set_password(password)
                request.user.save()
                login(request, request.user)

        if avatar:
            profile_obj.avatar = avatar
            profile_obj.save()

        request.user.save()

        if not error:
            success_message = "Cập nhật thông tin thành công."

    return render(request, "user_profile.html", {
        "profile_obj": profile_obj,
        "my_reservations": my_reservations,
        "approved_count": approved_count,
        "rejected_count": rejected_count,
        "pending_count": pending_count,
        "notifications": notifications,
        "success_message": success_message,
        "error": error,
    })


# ====== FAVORITES ======
@login_required
def toggle_favorite(request, restaurant_id):
    if request.method != "POST":
        return JsonResponse({"success": False}, status=405)

    restaurant = get_object_or_404(Restaurant, id=restaurant_id, is_active=True)

    fav = FavoriteRestaurant.objects.filter(
        user=request.user,
        restaurant=restaurant,
    ).first()

    if fav:
        fav.delete()
        return JsonResponse({"success": True, "favorited": False})

    FavoriteRestaurant.objects.create(user=request.user, restaurant=restaurant)
    return JsonResponse({"success": True, "favorited": True})


@login_required
def my_favorites(request):
    favorites = FavoriteRestaurant.objects.filter(user=request.user).select_related(
        "restaurant", "restaurant__category"
    )
    return render(
        request,
        "my_favorites.html",
        {
            "favorites": favorites,
        },
    )

# ===== LANDING PAGE =====
from django.db.models import Count

def landing_page(request):
    hot_restaurants = Restaurant.objects.filter(is_active=True).order_by("-price_level")[:8]

    popular_restaurants = (
        Restaurant.objects.filter(is_active=True)
        .annotate(total_reservations=Count("reservation"))
        .order_by("-total_reservations", "-price_level")[:6]
    )

    total_restaurants = Restaurant.objects.filter(is_active=True).count()
    total_categories = Category.objects.count()
    total_tables = DiningTable.objects.count()
    total_reservations = Reservation.objects.count()

    top_rated_restaurants = (
    Restaurant.objects.filter(is_active=True)
    .annotate(avg_rating=Avg("reviews__rating"))
    .order_by("-avg_rating")[:6]
)

    return render(request, "landing.html", {
        "hot_restaurants": hot_restaurants,
        "popular_restaurants": popular_restaurants,
        "total_restaurants": total_restaurants,
        "total_categories": total_categories,
        "total_tables": total_tables,
        "total_reservations": total_reservations,
        "top_rated_restaurants": top_rated_restaurants,
    })



# ====== STATS ======
def stats_by_category(request):
    data = (
        Restaurant.objects.values("category__name")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    return JsonResponse(
        {
            "data": [
                {
                    "category": d["category__name"] or "Chưa phân loại",
                    "total": d["total"],
                }
                for d in data
            ]
        }
    )

def clean_number(value):
    if not value:
        return 0
    return int(str(value).replace(".", "").replace(",", "").strip())

@login_required
def hang_hoa(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    products = Product.objects.select_related("restaurant", "category").all().order_by("-created_at")
    restaurants = Restaurant.objects.filter(is_active=True).order_by("name")
    categories = Category.objects.all().order_by("name")

    if request.method == "POST":
        product_id = request.POST.get("product_id")
        product = Product.objects.get(id=product_id) if product_id else Product()

        product.code = request.POST.get("code", "")
        product.name = request.POST.get("name", "")
        product.food_type = request.POST.get("food_type", "mon_an")
        product.category_id = request.POST.get("category") or None
        product.restaurant_id = request.POST.get("restaurant") or None
        product.status = request.POST.get("status", "active")
        product.price = clean_number(request.POST.get("price"))
        product.stock = clean_number(request.POST.get("stock"))
        product.min_stock = clean_number(request.POST.get("min_stock"))
        product.max_stock = clean_number(request.POST.get("max_stock"))
        product.description = request.POST.get("description", "")

        image = request.FILES.get("image")
        if image:
            product.image = image

        product.save()
        return redirect("hang_hoa")

    return render(request, "hang_hoa.html", {
        "products": products,
        "restaurants": restaurants,
        "categories": categories,
        "total_products": products.count(),
        "active_products": products.filter(status="active").count(),
        "out_products": products.filter(stock=0).count(),
    })

@login_required
def delete_product(request, product_id):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    product = get_object_or_404(Product, id=product_id)
    product.delete()
    return redirect("hang_hoa")

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@login_required
def update_price(request):
    if request.method == "POST":
        product_id = request.POST.get("id")
        price = request.POST.get("price")

        try:
            product = Product.objects.get(id=product_id)
            product.price = int(price)
            product.save()
            return JsonResponse({"success": True})
        except:
            return JsonResponse({"success": False})
        
@login_required
def update_stock(request, product_id):
    if request.method == "POST":
        product = get_object_or_404(Product, id=product_id)

        try:
            stock = int(request.POST.get("stock", 0))
            product.stock = stock

            # tự cập nhật trạng thái
            if stock == 0:
                product.status = "out_of_stock"
            elif stock < product.min_stock:
                product.status = "inactive"
            else:
                product.status = "active"

            product.save()

            return JsonResponse({"success": True})

        except:
            return JsonResponse({"success": False})

    return JsonResponse({"success": False})

@login_required
def export_products(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    response["Content-Disposition"] = 'attachment; filename="hang_hoa.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Ma hang",
        "Ten hang hoa",
        "Loai thuc don",
        "Nhom hang",
        "Gia ban",
        "Ton kho",
        "Trang thai",
    ])

    products = Product.objects.select_related("category").all()

    for p in products:
        writer.writerow([
            p.code,
            p.name,
            p.get_food_type_display(),
            p.category.name if p.category else "Chua phan loai",
            p.price,
            p.stock,
            p.get_status_display(),
        ])

    return response

@login_required
def export_revenue_excel(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    wb = Workbook()
    ws = wb.active
    ws.title = "Doanh thu"

    ws.merge_cells("A1:F1")
    ws["A1"] = "BÁO CÁO DOANH THU"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["STT", "Khách hàng", "Nhà hàng", "Số khách", "Thời gian", "Doanh thu"]
    ws.append(headers)

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F4BFD4")
        cell.alignment = Alignment(horizontal="center")

    reservations = Reservation.objects.filter(
        status="confirmed"
    ).select_related("restaurant").order_by("-booking_time")

    total = 0

    for index, r in enumerate(reservations, start=1):
        revenue = r.guests * r.restaurant.price_level * 100000
        total += revenue

        ws.append([
            index,
            r.customer_name,
            r.restaurant.name,
            r.guests,
            r.booking_time.strftime("%d/%m/%Y %H:%M"),
            revenue,
        ])

    ws.append([])
    ws.append(["", "", "", "", "Tổng doanh thu", total])

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 4

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="bao_cao_doanh_thu.xlsx"'

    wb.save(response)
    return response

@login_required
def nhan_vien(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    employees = Employee.objects.select_related("user", "restaurant").all().order_by("-created_at")
    restaurants = Restaurant.objects.filter(is_active=True).order_by("name")

    q = request.GET.get("q", "").strip()
    position = request.GET.get("position", "")
    status = request.GET.get("status", "")
    restaurant_id = request.GET.get("restaurant", "")

    if q:
        employees = employees.filter(user__username__icontains=q)

    if position:
        employees = employees.filter(position=position)

    if status:
        employees = employees.filter(status=status)

    if restaurant_id:
        employees = employees.filter(restaurant_id=restaurant_id)

    if request.method == "POST":
        employee_id = request.POST.get("employee_id")
        full_name = request.POST.get("full_name", "").strip()
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "").strip()
        phone = request.POST.get("phone", "").strip()
        position_post = request.POST.get("position", "waiter")
        status_post = request.POST.get("status", "active")
        restaurant_post = request.POST.get("restaurant") or None
        avatar = request.FILES.get("avatar")
        bonus_post = request.POST.get("bonus") or 0
        penalty_post = request.POST.get("penalty") or 0

        if employee_id:
            emp = get_object_or_404(Employee, id=employee_id)
            user = emp.user
        else:
            if User.objects.filter(username=username).exists():
                messages.error(request, "Tên đăng nhập đã tồn tại")
                return redirect("nhan_vien")

            user = User.objects.create_user(
                username=username,
                email=email,
                password=password or "123456"
            )
            user.is_staff = True
            user.save()
            emp = Employee(user=user)

        user.username = username
        user.email = email

        if password:
            user.set_password(password)

        if full_name:
            parts = full_name.split()
            if len(parts) == 1:
                user.first_name = parts[0]
                user.last_name = ""
            else:
                user.first_name = " ".join(parts[:-1])
                user.last_name = parts[-1]

        user.save()

        emp.phone = phone
        emp.position = position_post
        emp.status = status_post
        emp.restaurant_id = restaurant_post
        emp.bonus = int(bonus_post)
        emp.penalty = int(penalty_post)

        if avatar:
            emp.avatar = avatar

        emp.save()
        return redirect("nhan_vien")

    shifts = WorkShift.objects.all().order_by("start_time")
    attendances = Attendance.objects.select_related("employee", "shift").order_by("-date")
    performances = EmployeePerformance.objects.select_related("employee").order_by("-year", "-month")

    month = int(request.GET.get("salary_month") or timezone.now().month)
    year = int(request.GET.get("salary_year") or timezone.now().year)

    salary_data = []
    for e in employees:
        employee_attendances = Attendance.objects.filter(
            employee=e,
            date__month=month,
            date__year=year,
            check_in__isnull=False,
            check_out__isnull=False
        )

        total_hours = 0
        overtime_hours = 0

        for a in employee_attendances:
            check_in = datetime.combine(a.date, a.check_in)
            check_out = datetime.combine(a.date, a.check_out)

            hours = (check_out - check_in).seconds / 3600
            total_hours += hours

            if hours > 8:
                overtime_hours += hours - 8

        if e.position == "manager":
            hourly_rate = 40000
            base_salary = total_hours * hourly_rate
        elif e.position == "cashier":
            hourly_rate = 27000
            base_salary = total_hours * hourly_rate
        elif e.position == "waiter":
            hourly_rate = 22000
            base_salary = total_hours * hourly_rate
        elif e.position == "kitchen":
            hourly_rate = 35000
            base_salary = total_hours * hourly_rate
        elif e.position == "shipper":
            hourly_rate = 0
            base_salary = 7000000
        else:
            hourly_rate = 0
            base_salary = 0

        overtime_salary = overtime_hours * hourly_rate * 1.5
        bonus = e.bonus
        penalty = e.penalty
        final_salary = base_salary + overtime_salary + bonus - penalty

        salary_data.append({
            "employee": e,
            "month": month,
            "year": year,
            "hourly_rate": int(hourly_rate),
            "total_hours": round(total_hours, 1),
            "overtime_hours": round(overtime_hours, 1),
            "base_salary": int(base_salary),
            "overtime_salary": int(overtime_salary),
            "bonus": int(bonus),
            "penalty": int(penalty),
            "final_salary": int(final_salary),
        })

    return render(request, "nhan_vien.html", {
        "employees": employees,
        "restaurants": restaurants,
        "total_employees": Employee.objects.count(),
        "active_employees": Employee.objects.filter(status="active").count(),
        "inactive_employees": Employee.objects.filter(status="inactive").count(),
        "shifts": shifts,
        "attendances": attendances,
        "performances": performances,
        "salary_data": salary_data,
        "salary_month": month,
        "salary_year": year,
    })

@login_required
def delete_employee(request, employee_id):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    emp = get_object_or_404(Employee, id=employee_id)
    user = emp.user
    emp.delete()
    user.delete()

    return redirect("nhan_vien")

from datetime import datetime
from django.utils import timezone

@login_required
def check_in_employee(request, employee_id):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    employee = get_object_or_404(Employee, id=employee_id)
    today = timezone.localdate()
    now_time = timezone.localtime().time()

    # Tự chọn ca theo giờ hiện tại
    shift = WorkShift.objects.filter(
        start_time__lte=now_time,
        end_time__gte=now_time
    ).first()

    # Nếu không nằm trong ca nào thì lấy ca gần nhất
    if not shift:
        shift = WorkShift.objects.order_by("start_time").first()

    attendance, created = Attendance.objects.get_or_create(
        employee=employee,
        date=today,
        defaults={
            "shift": shift,
            "check_in": now_time,
        }
    )

    attendance.shift = shift

    if not attendance.check_in:
        attendance.check_in = now_time

    # Kiểm tra đi trễ
    if shift and now_time > shift.start_time:
        attendance.status = "late"
    else:
        attendance.status = "on_time"

    attendance.save()
    return redirect("nhan_vien")

    if a.status == "late":
        penalty += 20000
    elif a.status == "absent":
        penalty += 100000
    elif a.status == "left_early":
        penalty += 30000


@login_required
def check_out_employee(request, employee_id):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    employee = get_object_or_404(Employee, id=employee_id)
    today = timezone.localdate()
    now_time = timezone.localtime().time()

    attendance = Attendance.objects.filter(
        employee=employee,
        date=today
    ).first()

    if attendance:
        attendance.check_out = now_time
        attendance.save()

    return redirect("nhan_vien")

@login_required
def export_salary_excel(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    month = int(request.GET.get("month") or timezone.now().month)
    year = int(request.GET.get("year") or timezone.now().year)

    employees = Employee.objects.select_related("user", "restaurant").all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bang luong"

    ws.append([
        "STT", "Nhân viên", "Chức vụ", "Tháng", "Năm",
        "Tổng giờ", "Giờ OT", "Lương giờ",
        "Lương gốc", "Lương OT", "Thưởng", "Thực nhận"
    ])

    for index, e in enumerate(employees, start=1):
        attendances = Attendance.objects.filter(
            employee=e,
            date__month=month,
            date__year=year,
            check_in__isnull=False,
            check_out__isnull=False
        )

        total_hours = 0
        overtime_hours = 0

        for a in attendances:
            check_in = datetime.combine(a.date, a.check_in)
            check_out = datetime.combine(a.date, a.check_out)
            hours = (check_out - check_in).seconds / 3600

            total_hours += hours
            if hours > 8:
                overtime_hours += hours - 8

        if e.position == "manager":
            hourly_rate = 40000
            base_salary = total_hours * hourly_rate
        elif e.position == "cashier":
            hourly_rate = 27000
            base_salary = total_hours * hourly_rate
        elif e.position == "waiter":
            hourly_rate = 22000
            base_salary = total_hours * hourly_rate
        elif e.position == "kitchen":
            hourly_rate = 35000
            base_salary = total_hours * hourly_rate
        elif e.position == "shipper":
            hourly_rate = 0
            base_salary = 7000000
        else:
            hourly_rate = 0
            base_salary = 0

        overtime_salary = overtime_hours * hourly_rate * 1.5
        bonus = 300000 if total_hours >= 160 else 0
        final_salary = base_salary + overtime_salary + bonus

        ws.append([
            index,
            e.user.username,
            e.get_position_display(),
            month,
            year,
            round(total_hours, 1),
            round(overtime_hours, 1),
            int(hourly_rate),
            int(base_salary),
            int(overtime_salary),
            int(bonus),
            int(final_salary),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="bang_luong_{month}_{year}.xlsx"'
    wb.save(response)
    return response

@login_required
def update_employee_performance(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    month = timezone.now().month
    year = timezone.now().year

    employees = Employee.objects.all()

    for e in employees:
        attendances = Attendance.objects.filter(
            employee=e,
            date__month=month,
            date__year=year
        )

        total_work_days = attendances.filter(check_in__isnull=False).count()
        late_days = attendances.filter(status="late").count()
        absent_days = attendances.filter(status="absent").count()

        score = 100
        score -= late_days * 5
        score -= absent_days * 15

        if score < 0:
            score = 0

        EmployeePerformance.objects.update_or_create(
            employee=e,
            month=month,
            year=year,
            defaults={
                "total_work_days": total_work_days,
                "late_days": late_days,
                "absent_days": absent_days,
                "score": score,
            }
        )

    return redirect("nhan_vien")

@login_required
def phong_ban(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    restaurants = Restaurant.objects.filter(is_active=True).order_by("name")
    tables = DiningTable.objects.select_related("restaurant").all().order_by("restaurant__name", "code")

    selected_date = request.GET.get("date") or timezone.now().date().isoformat()
    restaurant_id = request.GET.get("restaurant")
    table_type = request.GET.get("table_type")
    status = request.GET.get("status")
    q = request.GET.get("q", "").strip()

    reservations = Reservation.objects.select_related("restaurant").filter(
        booking_time__date=selected_date
    ).order_by("booking_time")

    if restaurant_id:
        reservations = reservations.filter(restaurant_id=restaurant_id)
        tables = tables.filter(restaurant_id=restaurant_id)

    if table_type:
        reservations = reservations.filter(booking_type=table_type)
        tables = tables.filter(table_type=table_type)

    if status:
        reservations = reservations.filter(status=status)

    if q:
        reservations = reservations.filter(customer_name__icontains=q)

    time_slots = [
        "06:30", "07:00", "07:30", "08:00", "08:30",
        "09:00", "09:30", "10:00", "10:30",
        "11:00", "11:30", "12:00", "12:30",
        "13:00", "13:30", "14:00", "14:30",
        "15:00", "15:30", "16:00", "16:30",
        "17:00", "17:30", "18:00", "18:30",
        "20:00", "20:30", "21:00", "21:30",
        "22:00", "22:30", "23:00",
    ]

    slot_rows = []

    show_restaurants = restaurants
    if restaurant_id:
        show_restaurants = restaurants.filter(id=restaurant_id)

    for res in show_restaurants:
        row = {
            "restaurant": res,
            "slots": []
        }

        restaurant_reservations = reservations.filter(restaurant=res)

        for slot in time_slots:
            bookings = []

            for r in restaurant_reservations:
                if r.status == "confirmed" and r.booking_time.strftime("%H:%M") == slot:
                    bookings.append(r)

            row["slots"].append({
                "time": slot,
                "bookings": bookings
            })

        slot_rows.append(row)

    return render(request, "phong_ban.html", {
        "restaurants": restaurants,
        "tables": tables,
        "reservations": reservations,
        "selected_date": selected_date,
        "slot_rows": slot_rows,
        "time_slots": time_slots,

        "total_tables": tables.count(),
        "available_tables": tables.filter(status="available").count(),
        "reserved_tables": tables.filter(status="reserved").count(),
        "occupied_tables": tables.filter(status="occupied").count(),
        "cleaning_tables": tables.filter(status="cleaning").count(),
    })

@login_required
def approve_booking_from_table(request, reservation_id):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    reservation = get_object_or_404(Reservation, id=reservation_id)

    if reservation.status == "pending":
        available_table = DiningTable.objects.filter(
            restaurant=reservation.restaurant,
            table_type=reservation.booking_type,
            status="available"
        ).first()

        if available_table:
            available_table.status = "reserved"
            available_table.save()

            reservation.status = "confirmed"
            reservation.save()

    return redirect("phong_ban")

@login_required
def save_table(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    if request.method == "POST":
        table_id = request.POST.get("table_id")
        table = get_object_or_404(DiningTable, id=table_id) if table_id else DiningTable()

        table.code = request.POST.get("code")
        table.restaurant_id = request.POST.get("restaurant")
        table.table_type = request.POST.get("table_type")
        table.capacity = int(request.POST.get("capacity") or 0)
        table.status = request.POST.get("status")
        table.save()

    return redirect("phong_ban")


@login_required
def delete_table(request, table_id):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    table = get_object_or_404(DiningTable, id=table_id)
    table.delete()
    return redirect("phong_ban")


@login_required
def change_table_status(request, table_id, status):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    table = get_object_or_404(DiningTable, id=table_id)

    if status in ["available", "reserved", "occupied", "cleaning"]:
        table.status = status
        table.save()

    return redirect("phong_ban")

reservations = TableReservation.objects.all()

for r in reservations:

    if (
        r.status == "confirmed"
        and timezone.now() > r.booking_time + timedelta(minutes=15)
    ):
        r.status = "late"
        r.save()

@login_required
def bao_cao(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    today = timezone.now().date()

    start_date = request.GET.get("start_date") or today.replace(day=1).isoformat()
    end_date = request.GET.get("end_date") or today.isoformat()

    reservations = Reservation.objects.select_related("restaurant").filter(
        booking_time__date__gte=start_date,
        booking_time__date__lte=end_date
    )

    confirmed = reservations.filter(status="confirmed")
    pending = reservations.filter(status="pending")
    cancelled = reservations.filter(status="cancelled")
    late = reservations.filter(status="late")

    total_reservations = reservations.count()
    confirmed_count = confirmed.count()
    pending_count = pending.count()
    cancelled_count = cancelled.count()
    late_count = late.count()

    total_guests = confirmed.aggregate(total=Sum("guests"))["total"] or 0

    total_revenue = 0
    for r in confirmed:
        total_revenue += r.guests * r.restaurant.price_level * 100000

    daily_revenue = {}
    for r in confirmed:
        day = r.booking_time.strftime("%d/%m")
        amount = r.guests * r.restaurant.price_level * 100000
        daily_revenue[day] = daily_revenue.get(day, 0) + amount

    revenue_labels = list(daily_revenue.keys())
    revenue_values = list(daily_revenue.values())

    top_restaurants = (
        confirmed.values("restaurant__name")
        .annotate(total=Count("id"))
        .order_by("-total")[:5]
    )

    popular_hours = {}
    for r in reservations:
        hour = r.booking_time.strftime("%H:%M")
        popular_hours[hour] = popular_hours.get(hour, 0) + 1

    busiest_hour = "-"
    if popular_hours:
        busiest_hour = max(popular_hours, key=popular_hours.get)

    employees = Employee.objects.select_related("user").all()
    total_employees = employees.count()

    return render(request, "bao_cao.html", {
        "start_date": start_date,
        "end_date": end_date,

        "total_revenue": total_revenue,
        "total_reservations": total_reservations,
        "confirmed_count": confirmed_count,
        "pending_count": pending_count,
        "cancelled_count": cancelled_count,
        "late_count": late_count,
        "total_guests": total_guests,
        "busiest_hour": busiest_hour,
        "total_employees": total_employees,

        "revenue_labels": revenue_labels,
        "revenue_values": revenue_values,
        "top_restaurants": top_restaurants,
    })

@login_required
def export_bao_cao_excel(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Không có quyền")

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    reservations = Reservation.objects.select_related("restaurant").all()

    if start_date:
        reservations = reservations.filter(booking_time__date__gte=start_date)

    if end_date:
        reservations = reservations.filter(booking_time__date__lte=end_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bao cao dat ban"

    ws.append([
        "Ngày đặt",
        "Giờ đặt",
        "Người đặt",
        "Số điện thoại",
        "Nhà hàng",
        "Số khách",
        "Loại bàn",
        "Trạng thái",
        "Doanh thu"
    ])

    for r in reservations:
        revenue = 0
        if r.status == "confirmed":
            revenue = r.guests * r.restaurant.price_level * 100000

        ws.append([
            r.booking_time.strftime("%d/%m/%Y"),
            r.booking_time.strftime("%H:%M"),
            r.customer_name,
            r.customer_phone,
            r.restaurant.name,
            r.guests,
            r.get_booking_type_display(),
            r.get_status_display(),
            revenue
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="bao_cao_dat_ban.xlsx"'

    wb.save(response)
    return response